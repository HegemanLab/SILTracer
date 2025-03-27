# -*- coding: utf-8 -*-
"""
Created on Thu Feb  6 08:10:29 2025

@author: Clinton Curry, Cohen Lab - University of Minnesota
"""



import os
import time
import glob
import csv
from pprint import pprint
import PySirius
from PySirius import SiriusSDK
from PySirius.models.lcms_submission_parameters import LcmsSubmissionParameters
from PySirius.models.account_credentials import AccountCredentials
from PySirius.rest import ApiException
from PySirius.models.aligned_feature import AlignedFeature
from PySirius.models.aligned_feature_opt_field import AlignedFeatureOptField
from PySirius.models.structure_candidate_formula import StructureCandidateFormula
from PySirius.models.structure_candidate_opt_field import StructureCandidateOptField
from PySirius.models.formula_candidate import FormulaCandidate
from PySirius.models.formula_candidate_opt_field import FormulaCandidateOptField
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd


def load_config(script_dir):
    """
    Load configuration parameters from Sirius_config.csv located in the same directory as this script.
    Expected keys: username, password, adduct_choices, zodiac_option, project_name, ppm_error, mass_accuracy.
    """
    
    config_file = os.path.join(script_dir, "Sirius_config.csv")
    config = {}
    try:
        with open(config_file, mode='r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if len(row) < 2:
                    continue
                key = row[0].strip().lower()
                value = row[1].strip()
                config[key] = value
    except FileNotFoundError:
        print(f"Configuration file not found: {config_file}")
    return config


def login_to_sirius_web(config):
    """
    Log in to the SIRIUS web service using credentials from config and return an authenticated API instance.
    """
    
    # Start API instance and wait for port file to be created
    sdk = SiriusSDK()
    api = sdk.attach_or_start_sirius(headless=True)
    time.sleep(5)
    
    # Find Sirius file directory for port file
    sirius_dir = os.path.expanduser("~/.sirius")
    # Locate port file
    port_files = glob.glob(os.path.join(sirius_dir, "*.port"))
    
    # Handle no port file located error
    if not port_files:
        print("No .port file found in the SIRIUS directory.")
        return None

    # Get path of port file
    port_file_path = port_files[0]
    # Open port file and retrieve port number
    with open(port_file_path, 'r') as port_file:
        port = int(port_file.read().strip())
    
    print(f"Using port: {port}")
    
    # Set API instance to use local host and determined port number
    configuration = PySirius.Configuration(
        host=f"http://localhost:{port}"
    )
    api.api_client.configuration = configuration

    # # Set login credentials from config
    credentials = PySirius.AccountCredentials().from_dict({
        "username": config.get("username", ""),
        "password": config.get("password", "")
    })

    # Attempt login
    try:
        api.account().login(True, credentials)
        print("Login successful.")
        return api
    except Exception as e:
        print(f"Exception when calling LoginAndAccountApi->login: {e}")
        return None


def export_features_to_csv(api, project_id, output_file_path, allowed_adducts=None):
    """
    Fetch feature data from SIRIUS and write selected columns to CSV file.
    """
    # If no allowed adducts are provided, use a default list.
    if allowed_adducts is None:
        allowed_adducts = ["[M + H]+", "[M + Na]+", "[M - H2O + H]+"]
    
    try:
        # Create an API instance for FeaturesApi
        features_api = PySirius.FeaturesApi(api.api_client)
        opt_fields = ["topAnnotations"]
        # Fetch aligned features
        aligned_features = features_api.get_aligned_features(project_id, opt_fields=opt_fields)
        
        # Write the features and their structures to a CSV file
        with open(output_file_path, mode='w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)
            header = [
                "Feature ID", "Feature Mass", "Retention Time", "Quality",
                "Structure Rank", "SMILES", "CSI Score", "Structure Adduct",
                "Molecular Formula", "SIRIUS Score", "Zodiac Score",
                "Formula ID", "Structure DB Info"
            ]
            csv_writer.writerow(header)

            # Loop through aligned features
            for feature in aligned_features:
                # Check if feature has both MS1 and MS/MS data. If not, skip
                if not (getattr(feature, 'has_ms1', False) and getattr(feature, 'has_ms_ms', False)):
                    continue
                    
                # Extract feature details
                feature_id = getattr(feature, 'aligned_feature_id', "N/A")
                feature_mass = getattr(feature, 'ion_mass', "N/A")
                feature_rt = getattr(feature, 'rt_apex_seconds', "N/A")
                quality = getattr(feature.quality, 'value', "N/A")

                try:
                    opt_fields = ['dbLinks']
                    # Fetch structure candidates for the feature
                    structure_candidates = features_api.get_structure_candidates(
                        project_id, feature_id, opt_fields=opt_fields
                    )
                    
                    # Filter structure candidates based on allowed adducts.
                    filtered_candidates = [
                        structure for structure in structure_candidates 
                        if getattr(structure, 'adduct', "N/A") in allowed_adducts
                    ]
                    
                    # Limit to the top 5 filtered structures
                    for structure in filtered_candidates[:5]:
                        # Extract structure details
                        rank = getattr(structure, 'rank', "N/A")
                        smiles = getattr(structure, 'smiles', "N/A")
                        csi_score = getattr(structure, 'csi_score', "N/A")
                        molecular_formula = getattr(structure, 'molecular_formula', "N/A")
                        adduct = getattr(structure, 'adduct', "N/A")
                        formula_id = getattr(structure, 'formula_id', "N/A")
                        db_links_raw = getattr(structure, 'db_links', [])
                        db_links = ", ".join([f"{link.name}:{link.id}" for link in db_links_raw])

                        sirius_score = "N/A"
                        zodiac_score = "N/A"

                        # Fetch formula candidate data if formula_id exists
                        if formula_id != "N/A":
                            try:
                                formula_candidate = features_api.get_formula_candidate(
                                    project_id, feature_id, formula_id
                                )
                                sirius_score = getattr(formula_candidate, 'sirius_score', "N/A")
                                zodiac_score = getattr(formula_candidate, 'zodiac_score', "N/A")
                            except Exception as fe:
                                print(f"Error fetching formula scores for feature {feature_id}: {fe}")

                        # Write one structure per row to CSV
                        csv_writer.writerow([
                            feature_id, feature_mass, feature_rt, quality,
                            rank, smiles, csi_score, adduct,
                            molecular_formula, sirius_score, zodiac_score,
                            formula_id, db_links
                        ])
                except Exception as e:
                    print(f"Error fetching structure candidates for feature {feature_id}: {e}")

        print(f"Feature-level data exported to: {output_file_path}")

    except Exception as e:
        print(f"Exception when exporting feature-level data: {e}")


def process_mzml_files(mzml_folder, project_name, project_dir, config):
    """
    Use authenticated API from login function to process .mzML files with SIRIUS 
    and run export function.
    """
    
    # Initialize variable for SDK
    sdk = SiriusSDK()
    # Call login function and return authenticated API instance
    api = login_to_sirius_web(config)
    if not api:
        print("Login failed. Exiting process.")
        return
    
    # Path to project file
    project_path = os.path.join(project_dir, f"{project_name}.sirius")

    # Create .sirius project space file
    ps_info = api.projects().create_project_space(project_name, project_path)
    print(f"Project space created: {ps_info.project_id}")

    # List for SIRIUS analysis input files
    input_files = []
    print("Searching for .mzML files in folder:", mzml_folder)
    # Fill list with input .mzML files from given folder path
    for file_name in os.listdir(mzml_folder):
        file_path = os.path.join(mzml_folder, file_name)
        if file_name.endswith(".mzML"):
            input_files.append(file_path)
            print(f"Found .mzML file: {file_path}")

    if not input_files:
        print("No .mzML files found in the folder.")
        return

    # set alignment parameters
    parameters = LcmsSubmissionParameters()
    
    # Run LCMS alignment
    try:
        api_response = api.projects().import_ms_run_data(ps_info.project_id, parameters, input_files=input_files)
        if api_response:
            print("Import successful.")
    except Exception as e:
        print(f"Exception when calling ProjectsApi->import_ms_run_data: {e}")
        return
    
    # Set job submission parameters
    job_config = api.jobs().get_default_job_config()
    job_config.spectra_search_params.enabled = False
    job_config.formula_id_params.enabled = True
    job_config.formula_id_params.mass_accuracy_ms2ppm = float(config.get("mass_accuracy", 5))
    job_config.fingerprint_prediction_params.enabled = True
    job_config.structure_db_search_params.enabled = True 
    job_config.canopus_params.enabled = True
    job_config.zodiac_params.enabled = config.get("zodiac_option", "false").lower() == "true"
    
    # Get the raw adduct choices string from the config
    adduct_str = config.get("adduct_choices", "")
    # Remove any leading or trailing double quotes
    adduct_str = adduct_str.strip('"')
    # Split on commas and strip extra whitespace
    adduct_list = [adduct.strip() for adduct in adduct_str.split(",") if adduct.strip()]
    # Set adduct parameter
    if adduct_list:
        job_config.enforced_adducts = adduct_list
    else:
        job_config.enforced_adducts = [None]
    
    print(f"Fallback adducts set: {job_config.fallback_adducts}")
    print(f"Enforced adducts set: {job_config.enforced_adducts}")
    
    # Record the start time of the job
    job_start_time = time.time()
    
    # Run analysis job
    job = api.jobs().start_job(project_id=ps_info.project_id, job_submission=job_config)
    print(f"Job {job.id} started...")

    # Print recurring message while job is running
    while True:
        job_status = api.jobs().get_job(ps_info.project_id, job.id).progress.state
        if job_status == 'DONE':
            print(f"Job {job.id} completed.")
            break
        else:
            print(f"Job {job.id} still running. Waiting for 10 seconds...")
            time.sleep(10)


    # Calculate, format and print the elapsed time
    job_end_time = time.time()
    elapsed_time = job_end_time - job_start_time
    elapsed_hours = int(elapsed_time // 3600)
    elapsed_minutes = int((elapsed_time % 3600) // 60)
    elapsed_seconds = int(elapsed_time % 60)
    print(f"Total job time: {elapsed_hours} hours, {elapsed_minutes} minutes, {elapsed_seconds} seconds")

    # Define output CSV file path
    csv_file_path = os.path.join(project_dir, f"{project_name}_summary.csv")
    # Call csv function to export data to csv
    export_features_to_csv(api, ps_info.project_id, csv_file_path, allowed_adducts=job_config.enforced_adducts)

    # Shutdown SIRIUS instance using sdk
    sdk.shutdown_sirius()
    print(f"SIRIUS instance shut down. Project saved to: {project_path}")


def extract_hyperlinks(excel_file, sheet_name=None):
    """
    Function for extracting hyperlinks from regression output for writing to matched output.
    """
    # Load workbook
    wb = openpyxl.load_workbook(excel_file, data_only=False)
    # Select worksheet
    ws = wb.active if sheet_name is None else wb[sheet_name]

    # Retrieve the headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    # Index hyperlink columns
    plot_link_idx = headers.index("Plot Hyperlink")
    formula_link_idx = headers.index("Formula Hyperlink")
    peak_link_idx = headers.index("Peak Hyperlink")
    
    # Lists to store hyperlinks
    plot_links = []
    formula_links = []
    peak_links = []
    
    # Iterate over each row to extract cell values
    for row in ws.iter_rows(min_row=2):
        plot_cell = row[plot_link_idx]
        formula_cell = row[formula_link_idx]
        peak_cell = row[peak_link_idx]
        
        # Check if the cell has a hyperlink. If it does, extract the hyperlink target. 
        # Otherwise, use the cell's text value.
        plot_link = plot_cell.hyperlink.target if plot_cell.hyperlink else plot_cell.value
        formula_link = formula_cell.hyperlink.target if formula_cell.hyperlink else formula_cell.value
        peak_link = peak_cell.hyperlink.target if peak_cell.hyperlink else peak_cell.value
        
        # Append data to lists
        plot_links.append(plot_link)
        formula_links.append(formula_link)
        peak_links.append(peak_link)
    
    # Return lists
    return plot_links, formula_links, peak_links


def find_matching_features(regression_file, sirius_file, mz_col1, rt_col1, mz_col2, rt_col2, 
                           ppm_tolerance, rt_tolerance,output_file, regression_columns_to_save):
    """
    Match features between the regression dataset and SIRIUS output using mass error and RT tolerances.
    Saves regression, SIRIUS, and matched output into one Excel file.
    """
    
    # Load the Regression Excel file into dataframe
    regression_df = pd.read_excel(regression_file)
    
    # Extract hyperlinks and update the regression dataframe
    plot_links, formula_links, peak_links = extract_hyperlinks(regression_file)
    regression_df["Plot Hyperlink"] = plot_links
    regression_df["Formula Hyperlink"] = formula_links
    regression_df["Peak Hyperlink"] = peak_links

    # Load SIRIUS output CSV into dataframe
    sirius_df = pd.read_csv(sirius_file)

    # Convert retention time from seconds to minutes
    sirius_df[rt_col2] = sirius_df[rt_col2] / 60

    # Drop Formula ID column from SIRIUS data
    if 'Formula ID' in sirius_df.columns:
        sirius_df = sirius_df.drop(columns=['Formula ID'])

    # Wrap SIRIUS_Feature ID in quotes to prevent scientific notation in Excel
    if 'Feature ID' in sirius_df.columns:
        sirius_df['Feature ID'] = sirius_df['Feature ID'].apply(lambda x: f'"{x}"')

    # Create list to store matching features
    matches = []

    # Iterate through each feature in Regression data
    for i, regression_row in regression_df.iterrows():
        # set measured mass from regression data
        measured_mass = regression_row[mz_col1]

        # Compute acceptable mass range for matching based on PPM tolerance
        sirius_df["Lower Bound"] = sirius_df[mz_col2] * (1 - (ppm_tolerance / 1_000_000))
        sirius_df["Upper Bound"] = sirius_df[mz_col2] * (1 + (ppm_tolerance / 1_000_000))

        # Filter SIRIUS data based on the computed PPM range & retention time tolerance
        matched_sirius_rows = sirius_df[
            (sirius_df["Lower Bound"] <= measured_mass) &
            (measured_mass <= sirius_df["Upper Bound"]) &
            (abs(regression_row[rt_col1] - sirius_df[rt_col2]) <= rt_tolerance)
        ]

        # If matches found, process them
        if not matched_sirius_rows.empty:
            # loop through each row in filtered SIRIUS rows
            for _, sirius_row in matched_sirius_rows.iterrows():
                # Create a dictionary for each match
                match = {}
                
                # Copy over specified columns from the regression row
                for col in regression_columns_to_save:
                    match[col] = regression_row[col]
                # Copy columns from SIRIUS row (except the temporary bounds columns),
                # prefixing with 'SIRIUS_'
                for col in sirius_df.columns:
                    if col not in ["Lower Bound", "Upper Bound"]:
                        match[f'SIRIUS_{col}'] = sirius_row[col]
                        
                # Calculate the ppm error between the measured mass and theoretical mass
                theoretical_mass = sirius_row[mz_col2]
                ppm_error = ((measured_mass - theoretical_mass) / theoretical_mass) * 1_000_000 if theoretical_mass != 0 else None
                match['Mass Error (ppm)'] = ppm_error
                # Calculate difference in retention times
                match['RT Difference'] = abs(regression_row[rt_col1] - sirius_row[rt_col2])
                
                # Append this match dictionary to the list of matches
                matches.append(match)
                
    # Convert the list of matches into a DataFrame
    matches_df = pd.DataFrame(matches)

    # Format numeric columns to stop scientific notation
    for col in ['Mass Error (ppm)', 'RT Difference']:
        matches_df[col] = matches_df[col].apply(lambda x: f"{x:.10f}")

    # Generate a list of regression columns to keep in their original order
    regression_cols = [col for col in regression_columns_to_save]
    
    # Identify the columns that came from the SIRIUS data
    sirius_cols = [col for col in matches_df.columns if col.startswith("SIRIUS_")]
    
    # Arrange columns in order
    ordered_columns = regression_cols + sirius_cols + ["Mass Error (ppm)", "RT Difference"]
    matches_df = matches_df[ordered_columns]

    # Remove temporary bound columns for writing to output file
    sirius_output_df = sirius_df.drop(columns=["Lower Bound", "Upper Bound"], errors='ignore')
    
    # Write the regression DataFrame, SIRIUS DataFrame, and matches to an Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        regression_df.to_excel(writer, sheet_name='Regression Data', index=False)
        sirius_output_df.to_excel(writer, sheet_name='SIRIUS Data', index=False)
        matches_df.to_excel(writer, sheet_name='Matches', index=False)
        
        # Dictionary to map sheets to dataframes
        sheets = {
            "Regression Data": regression_df,
            "SIRIUS Data": sirius_output_df,
            "Matches": matches_df
        }
        
        # Adjusts output file column width for readability. Based on character count.
        for sheet_name, df in sheets.items():
            worksheet = writer.sheets[sheet_name]
            for i, col in enumerate(df.columns, start=1):
                col_letter = get_column_letter(i)
                header_text = str(col) if col is not None else ''
                worksheet.column_dimensions[col_letter].width = len(header_text) + 2

    print(f"Matches saved to {output_file}")
    return matches_df


def main():
    """
    Orchestral function that coordinates and executes the workflow.
    """
    # Determine working directory, based on script location.
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Load configuration from Sirius_config.csv
    config = load_config(script_dir)
    if not config:
        print("Configuration could not be loaded. Exiting.")
        return

    # Determine dynamic paths based on script location
    mzml_folder = os.path.join(script_dir, "Data", "Raw Files_SIRIUS")
    project_dir = os.path.join(script_dir, "Output")
    regression_file = os.path.join(script_dir, "Output", "FinalOutput.xlsx")
    output_file = os.path.join(script_dir, "Output", "matched_output.xlsx")

    # Parameters
    project_name = config.get("project_name", "default_project")
    ppm_tolerance = float(config.get("ppm_error", 5))
    rt_tolerance = 0.25
    mz_col1 = 'mzUnlabeled'
    rt_col1 = 'rtUnlabeled'
    mz_col2 = 'Feature Mass'
    rt_col2 = 'Retention Time'
    regression_columns_to_save = [
        'Unlabeled_Labeled Features', 'CompleteScore', 'mzUnlabeled', 'mzLabeled',
        'rtUnlabeled', 'NumLabels', 'Tentative Matches', 'Tentative Matches PPM Error',
        'Plot Hyperlink', 'Formula Hyperlink', 'Peak Hyperlink'
    ]
    
    # Run 1st part of workflow - SIRIUS analysis and output
    process_mzml_files(mzml_folder, project_name, project_dir, config)

    # The SIRIUS CSV file is produced by the first part.
    sirius_file = os.path.join(project_dir, f"{project_name}_summary.csv")
    
    # Run 2nd part of workflow - feature matching and output
    matched_df = find_matching_features(
        regression_file, sirius_file, mz_col1, rt_col1, mz_col2, rt_col2,
        ppm_tolerance, rt_tolerance, output_file, regression_columns_to_save
    )

    print("Feature matching complete.")


# Run main()
if __name__ == '__main__':
    main()
